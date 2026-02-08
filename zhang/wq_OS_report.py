### From：https://www.codecopy.cn/post/ani5hi?pw=wq123

import requests
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

############# 配置 #################

username = ""
password = ""
start_date = "2026-02-01"  # alpha提交日期筛选
end_date = "2026-02-08"

####################################

def wait_get(url: str, sess: any, max_retries: int = 10) -> "Response":
    """
    发送带有重试机制的 GET 请求，直到成功或达到最大重试次数。
    此函数会根据服务器返回的 `Retry-After` 头信息进行等待，并在遇到 401 状态码时重新初始化配置。
    
    Args:
        url (str): 目标 URL。
        max_retries (int, optional): 最大重试次数，默认为 10。
    
    Returns:
        Response: 请求的响应对象。
    """
    retries = 0
    while retries < max_retries:
        while True:
            simulation_progress = sess.get(url)
            if simulation_progress.headers.get("Retry-After", 0) == 0:
                break
            time.sleep(float(simulation_progress.headers["Retry-After"]))
        if simulation_progress.status_code < 400:
            break
        else:
            time.sleep(2 ** retries)
            retries += 1
    return simulation_progress

def login():
    # Create a session to persistently store the headers
    s = requests.Session()
 
    # Save credentials into session
    s.auth = (username, password)
 
    # Send a POST request to the /authentication API
    response = s.post('https://api.worldquantbrain.com/authentication')
    print(response.content)
    return s  

def get_alphas(start_date, end_date):
    s = login()
    output = []
    # 3E large 3C less
    count = 0
    offset = 0

    while True:
        
        # 构建URL 
        url = f"https://api.worldquantbrain.com/users/self/alphas?limit=100&offset={offset}" \
            + f"&status!=UNSUBMITTED%1FIS-FAIL&dateSubmitted%3E={start_date}T05:00:00.000Z" \
            + f"&dateSubmitted%3C={end_date}T05:00:00.000Z&order=-dateSubmitted&hidden=false&type!=SUPER"
        
        try:
            response = wait_get(url, s)
            alpha_list = response.json().get("results", [])
            
            # 如果没有数据，退出循环
            if not alpha_list:
                print("查询结束，没有更多数据")
                break
            else:
                print(f"{len(alpha_list)}条数据开始处理...")
            
            # 处理数据
            for alpha in alpha_list:
                count += 1
                # 提取数据...
                # 使用.get()方法避免KeyError
                alpha_id = alpha.get("id")
                date_submit = alpha.get("dateSubmitted")[:10]
                region = alpha.get("settings", {}).get("region")
                delay = alpha.get("settings", {}).get("delay")
                tags = alpha.get("tags", [])
                
                # 提取pyramids
                pyramid_items = alpha.get("pyramidThemes", {}).get("pyramids", [])
                pyramids = [item.get("name", "").split("/")[-1] for item in pyramid_items]
                
                # 获取年化统计数据
                try:
                    yearly_url = f"https://api.worldquantbrain.com/alphas/{alpha_id}/recordsets/yearly-stats"
                    yearly_response = wait_get(yearly_url, s)
                    records = yearly_response.json().get("records", [])
                    
                    if len(records) == 13:
                        sharp2022 = records[10][6] 
                        sharp2023 = records[12][6] 
                    else:
                        sharp2022 = records[9][6] 
                        sharp2023 = records[11][6] 
                except Exception:
                    sharp2022 = None
                    sharp2023 = None
                
                temp = (alpha_id, date_submit, region, delay, tags, pyramids, sharp2022, sharp2023)
                print(f"第{count}条数据：{temp}")
                output.append(temp)
            
            # 增加偏移量
            offset += 100
            
        except Exception as e:
            print(f"查询偏移量 {offset} 时出错: {e}")

    print(f"总共查询到 {count} 条数据")

    return output


s = login()

alpha_infos = get_alphas(start_date, end_date)

df = pd.DataFrame(alpha_infos, columns=['Alpha ID', 'Submit Date', 'Region', 'Delay',
                                 'Tags', 'Pyramids', 'Sharpe 2022', 'Sharpe 2023'])
df['Tags'] = df['Tags'].astype(str).str.strip("[]").str.replace("'", "")
df['Pyramids'] = df['Pyramids'].astype(str).str.strip("[]").str.replace("'", "")

# 将Submit Date转换为datetime格式并移除时区信息，然后提取月份用于后续统计
df['Submit Date'] = pd.to_datetime(df['Submit Date'], utc=True).dt.tz_localize(None)
df['Month'] = df['Submit Date'].dt.strftime('%Y-%m')
df['Submit Date'] = df['Submit Date'].dt.date

file = 'alpha_OS_report.xlsx'
with pd.ExcelWriter(file, engine='openpyxl') as w:
    # 写入主数据，排除Month列
    df[['Alpha ID', 'Submit Date', 'Region', 'Delay', 'Tags', 'Pyramids', 'Sharpe 2022', 'Sharpe 2023']].to_excel(w, index=False, sheet_name='Alpha OS分析', startrow=3)  # 从第4行开始写入数据
    ws = w.sheets['Alpha OS分析']
    
    # 在第一行添加颜色说明
    ws['A1'] = "Sharpe 2023颜色填充说明："
    ws['A1'].font = Font(name='Arial', bold=True, size=12)
    
    # 第二行添加具体说明
    ws['A2'] = "红色: Sharpe 2023 < 1 | 绿色: Sharpe 2023 ≥ Sharpe 2022| 黄色: Sharpe 2022 > Sharpe 2023 ≥ 1"
    ws['A2'].font = Font(name='Arial', bold=True, size=11, color='000000')
    # 合并第二行单元格
    ws.merge_cells('A2:H2')
    
    # 第四行是实际表头（因为我们从第四行开始写入数据，所以第三行是空行）
    
    # ---------- 表头样式 ----------
    header_font = Font(name='Arial', bold=True, size=13, color='ffffff')
    header_fill = PatternFill('solid', fgColor='3C8A97')
    header_align = Alignment(horizontal='center', vertical='center')
    
    # 表头在第4行
    for cell in ws[4]:  # 第4行是表头
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[4].height = 30  # 表头行高

    # ---------- 列宽 ----------
    for k, v in dict(zip('ABCDEFGH', [20, 15, 10, 10, 25, 20, 15, 15])).items():
        ws.column_dimensions[k].width = v

    # ---------- 数据区字体 + 行高 ----------
    data_font = Font(name='Arial', size=11)
    data_align = Alignment(vertical='center')
    
    # 定义边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 数据从第5行开始（表头在第4行，数据从第5行开始）
    for row in range(5, len(df) + 5):
        ws.row_dimensions[row].height = 18
        for col in range(1, 9):  # A-H
            c = ws.cell(row, col)
            c.font = data_font
            c.alignment = data_align
            c.border = thin_border  # 添加边框

    # ---------- 给 Sharpe 2023 涂色 ----------
    red = PatternFill('solid', fgColor='FF9999')
    green = PatternFill('solid', fgColor='99FF99')
    yellow = PatternFill('solid', fgColor='FFFF99')

    # 数据从第5行开始，所以行索引需要调整
    for row in range(5, len(df) + 5):
        s2023 = df.at[row - 5, 'Sharpe 2023']
        s2022 = df.at[row - 5, 'Sharpe 2022']
        cell = ws[f'H{row}']

        if s2023 < 1:
            cell.fill = red
        elif s2023 >= 1 and s2023 > s2022:
            cell.fill = green
        elif s2023 >= 1 and s2023 < s2022:
            cell.fill = yellow

    # ---------- 在主数据底部添加四个维度的报表 ----------
    
    # 确定报表开始的行（数据从第5行开始，所以需要加上偏移量）
    report_start_row = len(df) + 7  # 5(数据开始行) + len(df) - 1 + 3(空行)
    
    # 1. Region维度报表
    ws[f'A{report_start_row}'] = "Region维度报表"
    ws[f'A{report_start_row}'].font = Font(name='Arial', bold=True, size=12)
    
    # 报表表头
    ws[f'A{report_start_row+1}'] = "Region"
    ws[f'B{report_start_row+1}'] = "avg(sharpe2023)"
    ws[f'C{report_start_row+1}'] = "count(2022 < 2023)"
    ws[f'D{report_start_row+1}'] = "count(1 < 2023 < 2022)"
    ws[f'E{report_start_row+1}'] = "count(2023 < 1)"
    
    # 计算Region维度的统计数据
    region_stats = []
    for region in df['Region'].unique():
        region_data = df[df['Region'] == region]
        avg_sharpe = region_data['Sharpe 2023'].mean()
        count_better = len(region_data[region_data['Sharpe 2023'] > region_data['Sharpe 2022']])
        count_less_than_one = len(region_data[region_data['Sharpe 2023'] < 1])
        count_middle = len(region_data[(region_data['Sharpe 2023'] > 1) & (region_data['Sharpe 2023'] < region_data['Sharpe 2022'])])
        region_stats.append((region, avg_sharpe, count_better, count_middle, count_less_than_one))
    
    # 写入Region统计数据
    for i, (region, avg_sharpe, count_better, count_middle, count_less_than_one) in enumerate(region_stats):
        row = report_start_row + 2 + i
        ws[f'A{row}'] = region
        ws[f'B{row}'] = round(avg_sharpe, 2)
        ws[f'C{row}'] = count_better
        ws[f'D{row}'] = count_middle
        ws[f'E{row}'] = count_less_than_one
    
    # 2. Tags维度报表 - 修改为是否含ppa(PowerPoolSelected)
    tags_start_row = report_start_row + len(region_stats) + 4
    ws[f'A{tags_start_row}'] = "Tags维度报表"
    ws[f'A{tags_start_row}'].font = Font(name='Arial', bold=True, size=12)
    
    # 报表表头
    ws[f'A{tags_start_row+1}'] = "PowerPoolSelected"
    ws[f'B{tags_start_row+1}'] = "avg(sharpe2023)"
    ws[f'C{tags_start_row+1}'] = "count(2022 < 2023)"
    ws[f'D{tags_start_row+1}'] = "count(1 < 2023 < 2022)"
    ws[f'E{tags_start_row+1}'] = "count(2023 < 1)"
    
    # 计算Tags维度的统计数据 - 按是否包含PowerPoolSelected分类
    tags_stats = []
    
    # 含ppa(PowerPoolSelected)
    ppa_data = df[df['Tags'].str.contains('PowerPoolSelected', na=False)]
    avg_sharpe_ppa = ppa_data['Sharpe 2023'].mean()
    count_better_ppa = len(ppa_data[ppa_data['Sharpe 2023'] > ppa_data['Sharpe 2022']])
    count_less_than_one_ppa = len(ppa_data[ppa_data['Sharpe 2023'] < 1])
    count_middle_ppa = len(ppa_data[(ppa_data['Sharpe 2023'] > 1) & (ppa_data['Sharpe 2023'] < ppa_data['Sharpe 2022'])])
    tags_stats.append(('含PowerPoolSelected', avg_sharpe_ppa, count_better_ppa, count_middle_ppa, count_less_than_one_ppa))
    
    # 不含ppa(PowerPoolSelected)
    no_ppa_data = df[~df['Tags'].str.contains('PowerPoolSelected', na=True)]
    avg_sharpe_no_ppa = no_ppa_data['Sharpe 2023'].mean()
    count_better_no_ppa = len(no_ppa_data[no_ppa_data['Sharpe 2023'] > no_ppa_data['Sharpe 2022']])
    count_less_than_one_no_ppa = len(no_ppa_data[no_ppa_data['Sharpe 2023'] < 1])
    count_middle_no_ppa = len(no_ppa_data[(no_ppa_data['Sharpe 2023'] > 1) & (no_ppa_data['Sharpe 2023'] < no_ppa_data['Sharpe 2022'])])
    tags_stats.append(('不含PowerPoolSelected', avg_sharpe_no_ppa, count_better_no_ppa, count_middle_no_ppa, count_less_than_one_no_ppa))
    
    # 写入Tags统计数据
    for i, (tag_type, avg_sharpe, count_better, count_middle, count_less_than_one) in enumerate(tags_stats):
        row = tags_start_row + 2 + i
        ws[f'A{row}'] = tag_type
        ws[f'B{row}'] = round(avg_sharpe, 2)
        ws[f'C{row}'] = count_better
        ws[f'D{row}'] = count_middle
        ws[f'E{row}'] = count_less_than_one
    
    # 3. Pyramid维度报表
    pyramid_start_row = tags_start_row + len(tags_stats) + 4
    ws[f'A{pyramid_start_row}'] = "Pyramid维度报表"
    ws[f'A{pyramid_start_row}'].font = Font(name='Arial', bold=True, size=12)
    
    # 报表表头
    ws[f'A{pyramid_start_row+1}'] = "Pyramid"
    ws[f'B{pyramid_start_row+1}'] = "avg(sharpe2023)"
    ws[f'C{pyramid_start_row+1}'] = "count(2022 < 2023)"
    ws[f'D{pyramid_start_row+1}'] = "count(1 < 2023 < 2022)"
    ws[f'E{pyramid_start_row+1}'] = "count(2023 < 1)"
    
    # 计算Pyramid维度的统计数据
    pyramid_stats = []
    # 拆分Pyramid列中的多个值
    pyramid_list = []
    for index, row in df.iterrows():
        pyramids = str(row['Pyramids']).split(',')
        for pyramid in pyramids:
            pyramid_clean = pyramid.strip()
            if pyramid_clean:
                pyramid_list.append({
                    'Pyramid': pyramid_clean,
                    'Sharpe 2022': row['Sharpe 2022'],
                    'Sharpe 2023': row['Sharpe 2023']
                })
    
    pyramid_df = pd.DataFrame(pyramid_list)
    
    # 按Pyramid类型分组统计
    for pyramid_type in pyramid_df['Pyramid'].unique():
        pyramid_data = pyramid_df[pyramid_df['Pyramid'] == pyramid_type]
        avg_sharpe = pyramid_data['Sharpe 2023'].mean()
        count_better = len(pyramid_data[pyramid_data['Sharpe 2023'] > pyramid_data['Sharpe 2022']])
        count_less_than_one = len(pyramid_data[pyramid_data['Sharpe 2023'] < 1])
        count_middle = len(pyramid_data[(pyramid_data['Sharpe 2023'] > 1) & (pyramid_data['Sharpe 2023'] < pyramid_data['Sharpe 2022'])])
        pyramid_stats.append((pyramid_type, avg_sharpe, count_better, count_middle, count_less_than_one))
    
    # 按avg(sharpe2023)降序排序
    pyramid_stats.sort(key=lambda x: x[1], reverse=True)
    
    # 写入Pyramid统计数据
    for i, (pyramid_type, avg_sharpe, count_better, count_middle, count_less_than_one) in enumerate(pyramid_stats):
        row = pyramid_start_row + 2 + i
        ws[f'A{row}'] = pyramid_type
        ws[f'B{row}'] = round(avg_sharpe, 2)
        ws[f'C{row}'] = count_better
        ws[f'D{row}'] = count_middle
        ws[f'E{row}'] = count_less_than_one
    
    # 4. 时间月度维度报表
    month_start_row = pyramid_start_row + len(pyramid_stats) + 4
    ws[f'A{month_start_row}'] = "时间月度维度报表"
    ws[f'A{month_start_row}'].font = Font(name='Arial', bold=True, size=12)
    
    # 报表表头
    ws[f'A{month_start_row+1}'] = "Month"
    ws[f'B{month_start_row+1}'] = "avg(sharpe2023)"
    ws[f'C{month_start_row+1}'] = "count(2022 < 2023)"
    ws[f'D{month_start_row+1}'] = "count(1 < 2023 < 2022)"
    ws[f'E{month_start_row+1}'] = "count(2023 < 1)"
    
    # 计算月度维度的统计数据
    month_stats = []
    for month in sorted(df['Month'].unique()):
        month_data = df[df['Month'] == month]
        avg_sharpe = month_data['Sharpe 2023'].mean()
        count_better = len(month_data[month_data['Sharpe 2023'] > month_data['Sharpe 2022']])
        count_less_than_one = len(month_data[month_data['Sharpe 2023'] < 1])
        count_middle = len(month_data[(month_data['Sharpe 2023'] > 1) & (month_data['Sharpe 2023'] < month_data['Sharpe 2022'])])
        month_stats.append((month, avg_sharpe, count_better, count_middle, count_less_than_one))
    
    # 写入月度统计数据
    for i, (month, avg_sharpe, count_better, count_middle, count_less_than_one) in enumerate(month_stats):
        row = month_start_row + 2 + i
        ws[f'A{row}'] = month
        ws[f'B{row}'] = round(avg_sharpe, 2)
        ws[f'C{row}'] = count_better
        ws[f'D{row}'] = count_middle
        ws[f'E{row}'] = count_less_than_one
    
    # 给报表表头添加样式
    report_header_font = Font(name='Arial', bold=True, size=11, color='ffffff')
    report_header_fill = PatternFill('solid', fgColor='5D8AA8')
    report_header_align = Alignment(horizontal='center', vertical='center')
    
    # 报表表头背景色和边框
    for start_row in [report_start_row+1, tags_start_row+1, pyramid_start_row+1, month_start_row+1]:
        for col in range(1, 6):  # A-E列
            cell = ws.cell(row=start_row, column=col)
            cell.font = report_header_font
            cell.fill = report_header_fill
            cell.alignment = report_header_align
            cell.border = thin_border
    
    # 给报表标题添加边框
    for title_row in [report_start_row, tags_start_row, pyramid_start_row, month_start_row]:
        cell = ws.cell(row=title_row, column=1)
        cell.border = Border(left=thin_border.left, right=thin_border.right, top=thin_border.top, bottom=thin_border.bottom)
    
    # 给报表数据区域添加边框和填充颜色
    for stats, start_row, num_rows in [
        (region_stats, report_start_row, len(region_stats)),
        (tags_stats, tags_start_row, len(tags_stats)),
        (pyramid_stats, pyramid_start_row, len(pyramid_stats)),
        (month_stats, month_start_row, len(month_stats))
    ]:
        for i in range(num_rows):
            row = start_row + 2 + i
            # 给所有数据单元格添加边框
            for col in range(1, 6):  # A-E列
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col in [2, 3, 4, 5]:  # B-E列，数字列居中对齐
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 给C列(count_better)填充绿色
            c_cell = ws.cell(row=row, column=3)  # C列
            c_cell.fill = green
            
            # 给D列(count_middle)填充黄色
            d_cell = ws.cell(row=row, column=4)  # D列
            d_cell.fill = yellow
            
            # 给E列(count_less_than_one)填充红色
            e_cell = ws.cell(row=row, column=5)  # E列
            e_cell.fill = red
    
    # 调整报表列宽
    for col in range(1, 6):  # A-E列
        col_letter = get_column_letter(col)
        max_length = 0
        # 查找所有报表区域的最大长度
        for row in range(report_start_row, month_start_row + len(month_stats) + 3):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 35)  # 最大宽度35
        ws.column_dimensions[col_letter].width = adjusted_width

print('已生成：', file)
